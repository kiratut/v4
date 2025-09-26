"""
Утилита проверки Excel-файла: считает строки/колонки, выводит заголовки и первые записи.
Сохраняет отчёт в utils/verify_excel_results.txt
"""

from pathlib import Path
import sys

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl не установлен")
    sys.exit(1)


def is_row_empty(values):
    return all((v is None or str(v).strip() == "") for v in values)


def main():
    if len(sys.argv) < 2:
        print("Usage: python utils/verify_excel.py <path_to_xlsx>")
        sys.exit(2)

    xlsx_path = Path(sys.argv[1])
    report_path = Path("utils/verify_excel_results.txt")

    lines = []
    lines.append(f"Файл: {xlsx_path}")
    if not xlsx_path.exists():
        lines.append("❌ Файл не найден")
    else:
        size = xlsx_path.stat().st_size
        lines.append(f"Размер: {size} байт")
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
            sheet = wb[wb.sheetnames[0]]
            lines.append(f"Лист: {sheet.title}")

            # Заголовки
            headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
            lines.append(f"Заголовки: {headers}")

            # Подсчёт непустых строк данных
            data_rows = 0
            preview = []
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if is_row_empty(row):
                    continue
                data_rows += 1
                if len(preview) < 3:
                    preview.append(list(row))

            lines.append(f"Строк данных (без заголовка): {data_rows}")
            lines.append("Первые строки:")
            for idx, r in enumerate(preview, 1):
                lines.append(f"  {idx}. {r}")

            wb.close()
        except Exception as e:
            lines.append(f"❌ Ошибка чтения: {e}")

    # Пишем отчёт
    report_path.write_text("\n".join(lines), encoding="utf-8")
    print("\n".join(lines))
    print(f"\n✅ Отчёт: {report_path}")


if __name__ == "__main__":
    main()
