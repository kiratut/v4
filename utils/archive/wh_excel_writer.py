"""
Модуль для записи данных в файл сценария Excel
"""

import pandas as pd
import logging
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from v4.wh_logger_config import format_worksheet
from v4.wh_global_params import GlobalParams
from typing import List, Dict, Any

logger = logging.getLogger(__name__)

def write_scenario_sheets(data_blocks: List[Dict[str, Dict[str, Any]]], template_path: str, scenario_path: str) -> bool:
    """
    Запись данных в файл сценария.
    
    Args:
        data_blocks: Список словарей {sheet_name: {
            'fields': {field_name: [values]},
            'use_template': bool, # использовать ли шаблонную строку для пропущенных значений
            'save_template': bool  # Нужно ли переносить шаблонную строку в сценарий
            'column_formats': {field_name: format}  # Форматы для колонок
        }}
    """
    logger.info("Начало записи данных в файл сценария")
    
    try:
        # Отладочная информация о типах данных
        logger.debug(f"Количество блоков данных: {len(data_blocks)}")
        for i, block in enumerate(data_blocks):
            logger.debug(f"Блок {i}: тип = {type(block)}, данные = {block if isinstance(block, dict) else str(block)[:100]}")
        
        # --- Записываем данные в листы Excel ---
        # Группируем данные по листам
        grouped_data = {}
        for i, block in enumerate(data_blocks):
            if not isinstance(block, dict):
                logger.error(f"Блок {i} не является словарем: тип = {type(block)}")
                continue
            for sheet_name, sheet_data in block.items():
                if sheet_name not in grouped_data:
                    grouped_data[sheet_name] = []
                grouped_data[sheet_name].append(sheet_data)
        
        # Открываем шаблон и файл сценария
        template_book = openpyxl.load_workbook(template_path)
        
        # Загружаем книгу сценария для прямого редактирования
        book = load_workbook(scenario_path)

        # Обрабатываем каждый лист
        for sheet_name, data_list in grouped_data.items():
            try:
                if sheet_name not in book.sheetnames:
                    logger.error(f"Лист '{sheet_name}' не найден в файле сценария. Пропускаем.")
                    continue

                ws = book[sheet_name]
                
                # Если save_template=False, очищаем лист перед записью
                if not data_list[0].get('save_template', True):
                    if ws.max_row > 1:
                        ws.delete_rows(2, ws.max_row - 1)
                        logger.info(f"Лист '{sheet_name}' очищен перед записью новых данных.")

                # Получаем заголовки и шаблон из файла шаблона
                if sheet_name not in template_book.sheetnames:
                    logger.error(f"Лист {sheet_name} не найден в шаблоне")
                    continue
                    
                ws_template = template_book[sheet_name]
                headers = [cell.value for cell in ws_template[1] if cell.value]
                
                # Строка 2 из шаблона используется для заполнения пропущенных полей
                template_row_for_defaults = [ws_template.cell(row=2, column=i+1).value for i in range(len(headers))]

                # Получаем форматы ячеек из шаблона
                cell_formats = {}
                for col_idx, header in enumerate(headers, start=1):
                    template_cell = ws_template.cell(row=2, column=col_idx)
                    cell_formats[header] = template_cell.number_format
                
                # Специальный формат для колонок "Start" и "End" на листе "Periods"
                if sheet_name == "Periods":
                    cell_formats["Start"] = 'M/d/yy HH:mm:ss'
                    cell_formats["End"] = 'M/d/yy HH:mm:ss'

                # Собираем все DataFrame для этого листа
                all_dfs = []
                for block_data in data_list:
                    fields_data = block_data.get('fields', {})
                    if not fields_data or not any(fields_data.values()):
                        continue
                    
                    rows_count = max(len(v) for v in fields_data.values() if v)
                    df_data = {}
                    use_template_defaults = block_data.get('use_template', True)

                    for header in headers:
                        if header in fields_data:
                            values = fields_data[header]
                            df_data[header] = values + [None] * (rows_count - len(values))
                        elif use_template_defaults:
                            template_value = template_row_for_defaults[headers.index(header)]
                            df_data[header] = [template_value] * rows_count
                        else:
                            df_data[header] = [None] * rows_count
                    
                    all_dfs.append(pd.DataFrame(df_data))

                if not all_dfs:
                    logger.info(f"Нет новых данных для записи на лист '{sheet_name}'. Пропускаем.")
                    continue
                
                combined_df = pd.concat(all_dfs, ignore_index=True)
                
                # Дописываем строки из DataFrame в конец листа
                rows = dataframe_to_rows(combined_df, index=False, header=False)
                
                for r_idx, row_values in enumerate(rows, start=ws.max_row + 1):
                    for c_idx, value in enumerate(row_values, start=1):
                        cell = ws.cell(row=r_idx, column=c_idx)
                        header = headers[c_idx - 1]
                        
                        # Обработка строк, начинающихся с =
                        if isinstance(value, str) and value.startswith("="):
                            cell._value = value
                            cell.data_type = 's'
                        else:
                            cell.value = value
                            
                        # Применяем формат из шаблона
                        if header in cell_formats:
                            cell.number_format = cell_formats[header]

                logger.info(f"На лист '{sheet_name}' добавлено {len(combined_df)} строк.")

            except Exception as e:
                logger.error(f"Ошибка при обработке листа {sheet_name}: {str(e)}", exc_info=True)
                continue
        
        # Перенумеровываем столбцы 'ID' на всех листах
        logger.info("Перенумерация столбцов 'ID' на всех листах...")
        GlobalParams.reset_id_counter()
        for ws_ids in book.worksheets:
            # Ищем колонку с точным заголовком 'ID'
            header_row = 1
            id_col_idx = None
            for col_idx in range(1, ws_ids.max_column + 1):
                if ws_ids.cell(row=header_row, column=col_idx).value == "ID":
                    id_col_idx = col_idx
                    break
            if id_col_idx is None:
                continue
            for row_idx in range(2, ws_ids.max_row + 1):
                if GlobalParams.ID_NUMBERING_MODE == 'PER_SHEET':
                    new_id_num = GlobalParams.get_next_id(ws_ids.title)
                else:
                    new_id_num = GlobalParams.get_next_id()
                ws_ids.cell(row=row_idx, column=id_col_idx).value = f"{GlobalParams.ID_PREFIX}{new_id_num}"
        logger.info("Перенумерация завершена.")

        # Форматируем все листы
        logger.info("Форматируем все листы")
        for ws in book.worksheets:
            format_worksheet(ws, None)  # Передаем None вместо writer

        # Сохраняем измененную книгу
        book.save(GlobalParams.SCENARIO_FILE_NAME)
        template_book.close()
        logger.info("Файл сценария успешно обновлен")
        return True
        
    except Exception as e:
        logger.error(f"Ошибка при записи файла сценария: {str(e)}", exc_info=True)
        return False
