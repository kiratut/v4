"""
Модуль настройки логирования и отладочного вывода.

Функции:
    setup_logging(log_file='debug.log', mode='w') -> logging.Logger:
        Настраивает и возвращает логгер с файловым и консольным выводом
        
    table_debug(data, file_name='table_debug.xlsx', sheet_names=None) -> None:
        Сохраняет DataFrame(ы) в Excel с автоматическим добавлением временных меток
        
Использование:
    logger = setup_logging()
    logger.info("Сообщение")
    
    table_debug([df1, df2], sheet_names=['Sheet1', 'Sheet2'])
"""

import logging
import pandas as pd
from pathlib import Path
from v4.wh_global_params import GlobalParams
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

def setup_logging(log_file: str = None) -> logging.Logger:
    """Настройка логирования"""
    # Создаем или получаем логгер
    logger = logging.getLogger('graph_layout')
    logger.setLevel(GlobalParams.LOG_LEVEL)
    
    # Очищаем существующие обработчики
    logger.handlers.clear()
    
    # Создаем упрощенный форматтер с временем
    formatter = logging.Formatter(
        '%(asctime)s %(levelname)s: %(message)s',
        datefmt='%H:%M:%S'
    )
    
    # Добавляем только файловый обработчик
    if log_file:
        # Перезаписываем файл при каждом запуске и пишем с BOM для корректного отображения в Windows
        file_handler = logging.FileHandler(log_file, encoding='utf-8-sig', mode='w')
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        
        # Отключаем propagation чтобы логи не шли в родительские обработчики
        logger.propagate = False
    
    return logger

def format_worksheet(worksheet, writer):
    """Форматирование листа Excel: автофильтр, закрепление строк, стили"""
    # Добавляем автофильтр и закрепляем первую строку
    if worksheet.dimensions:
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.freeze_panes = 'A2'
    
    # Создаем базовые стили
    base_alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
    header_font = Font(bold=True)
    
    # Применяем стили и автоподбор ширины
    max_lengths = [0] * (worksheet.max_column)
    
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value:
                # Применяем стили
                cell.alignment = base_alignment
                if cell.row == 1:  # Первая строка
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)  # Перенос текста только для заголовков
                
                # Обновляем максимальную длину для столбца
                try:
                    max_lengths[cell.column - 1] = max(
                        max_lengths[cell.column - 1], 
                        len(str(cell.value))
                    )
                except:
                    pass
    
    # Устанавливаем ширину столбцов, не более 10 символов
    for col, max_length in enumerate(max_lengths, 1):
        adjusted_width = min(max_length + 2, 10)  # Устанавливаем ширину не более 10
        worksheet.column_dimensions[get_column_letter(col)].width = adjusted_width

def convert_to_dataframe(data, filter_name='unspecified_filter', logger=None):
    """Преобразует различные табличные форматы в pandas DataFrame"""
    try:
        original_type = type(data).__name__
        
        # Добавляем проверку для None
        if data is None:
            if logger:
                logger.warning(f"Получены None данные для фильтра '{filter_name}'")
            return pd.DataFrame()
        
        if isinstance(data, pd.DataFrame):  # Уже DataFrame
            return data
        
        if isinstance(data, list) and all(isinstance(item, dict) for item in data):  # Список словарей
            df = pd.DataFrame(data)
            # Если есть общий ключ, добавляем его первой колонкой
            if all('key' in item for item in data):
                df.insert(0, 'key', [item['key'] for item in data])
        elif isinstance(data, dict):
            if all(isinstance(v, list) for v in data.values()):  # Словарь списков
                df = pd.DataFrame.from_dict(data, orient='index').stack().reset_index()
                df.columns = ['key', 'subkey', 'value']
            elif all(not isinstance(v, (list, dict)) for v in data.values()):  # Словарь скалярных значений
                df = pd.DataFrame(list(data.items()), columns=['key', 'value'])
                if logger:
                    logger.debug(f"Словарь скалярных значений преобразован в DataFrame с индексами")
            elif all(isinstance(v, tuple) for v in data.values()):  # Словарь кортежей
                df = pd.DataFrame.from_dict(data, orient='index')
                df.reset_index(inplace=True)
                df.columns = ['key', 'value1', 'value2']  # Переименовываем колонки
            else:  # Смешанный словарь
                df = pd.DataFrame([data])  # Преобразуем в список из одного словаря
                # Если есть ключ, добавляем его
                if 'key' in data:
                    df.insert(0, 'key', [data['key']])
        elif isinstance(data, str):  # Обработка строк
            df = pd.DataFrame([data], columns=['value'])
        elif hasattr(data, '__array__') or isinstance(data, (list, tuple)):  # numpy array или 2D список
            df = pd.DataFrame(data)
        else:  # Попытка стандартного преобразования
            df = pd.DataFrame(data)
        
        # Проверяем результат преобразования
        if df.empty:
            if logger:
                logger.warning(f"Пустой DataFrame после преобразования {original_type} для фильтра '{filter_name}'")
        else:
            if logger:
                logger.debug(f"Успешное преобразование {original_type} в DataFrame")
            
        return df
        
    except Exception as e:
        if logger:
            logger.error(f"Ошибка при преобразовании {original_type} в DataFrame для фильтра '{filter_name}': {str(e)}")
        return None
    
def table_debug(dataframes, file_name='table_debug.xlsx', sheet_names=None, mode='a', sheet_filter=None):
    """
    Сохраняет DataFrame(ы) в Excel с автоматическим добавлением временных меток
    Args:
        dataframes: список DataFrame для сохранения
        sheet_names: список имен листов
        file_name: имя файла Excel
        mode: режим записи ('w' или 'a')
        sheet_filter: список названий листов для фильтрации из Описание.xlsx
    """
    if sheet_names is None:
        sheet_names = [f'Sheet{i}' for i in range(len(dataframes))]
    
    excel_path = Path(file_name)
    logger = logging.getLogger('graph_layout')
    
    # Фильтрация колонок если указан sheet_filter
    filtered_dataframes = []
    if dataframes and sheet_filter and len(sheet_filter) == len(dataframes):
        try:
            # Читаем лист "каталог (2)" из файла описания
            desc_df = pd.read_excel(GlobalParams.DESCRIPTION_FILE, sheet_name=GlobalParams.CATALOG_SHEET)
            
            # Получаем все возможные param_name
            all_params = set(desc_df['param_name'].tolist())
            
            # Обрабатываем каждый DataFrame с соответствующим фильтром
            for df, filter_name in zip(dataframes, sheet_filter):
                
                if filter_name:
                    # Получаем param_name для текущего листа
                    sheet_params = set(desc_df[desc_df['Лист'] == filter_name]['param_name'].tolist())
                    
                    if sheet_params:
                        # Создаем список колонок для исключения
                        exclude_columns = all_params - sheet_params
                        
                        # Оставляем только те колонки, которые есть в DataFrame
                        valid_exclude = [col for col in exclude_columns if col in df.columns]
                        
                        # Получаем список колонок для сохранения
                        keep_columns = [col for col in df.columns if col not in valid_exclude]
                        
                        filtered_dataframes.append(df[keep_columns])
                        logger.debug(f"Для листа {filter_name}:")
                        logger.debug(f"  - Исключены колонки: {valid_exclude}")
                        logger.debug(f"  - Оставлены колонки: {keep_columns}")
                    else:
                        filtered_dataframes.append(df)
                        logger.debug(f"Лист {filter_name} не найден в каталоге (2)")
                else:
                    filtered_dataframes.append(df)
            
        except Exception as e:
            logger.error(f"Ошибка при чтении Описание.xlsx: {str(e)}")
            filtered_dataframes = dataframes
    else:
        filtered_dataframes = dataframes
    
    # Если файл не существует или явно указан режим 'w', создаем новый файл
    if not excel_path.exists() or mode == 'w':
        try:
            # Проверяем, если dataframes пустой
            if not dataframes:
                # Создаем новый файл с пустым листом
                with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
                    # Добавляем пустой DataFrame для создания листа
                    pd.DataFrame().to_excel(writer, sheet_name='EmptySheet', index=False)
                logger.debug(f"Создан файл {excel_path} с пустым листом")
                return
            
            with pd.ExcelWriter(excel_path, 
                              engine='openpyxl',
                              mode='w') as writer:
                for df, sheet_name in zip(filtered_dataframes, sheet_names):

                    # Проверяем и преобразуем данные
                    if not isinstance(df, pd.DataFrame):
                        # Добавляем значение по умолчанию для filter_name
                        current_filter = filter_name if 'filter_name' in locals() else 'unnamed_filter'
                        df = convert_to_dataframe(df, current_filter, logger)
                        if df is None:
                            continue

                    if isinstance(df, pd.DataFrame):
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        format_worksheet(writer.sheets[sheet_name], writer)
                    else:
                        logger.error(f"Не удалось сохранить таблицу '{sheet_name}' - неверный формат данных")
        except Exception as e:
            logger.error(f"Ошибка при сохранении таблиц в {excel_path}: {str(e)}")
    else:
        try:
            with pd.ExcelWriter(excel_path, 
                              engine='openpyxl',
                              mode='a',
                              if_sheet_exists='replace') as writer:
                for df, sheet_name in zip(filtered_dataframes, sheet_names):
                    # Проверяем и преобразуем данные
                    if not isinstance(df, pd.DataFrame):
                        # Добавляем значение по умолчанию для filter_name
                        current_filter = filter_name if 'filter_name' in locals() else 'unnamed_filter'
                        df = convert_to_dataframe(df, current_filter, logger)
                        if df is None:
                            continue

                    if isinstance(df, pd.DataFrame):
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        format_worksheet(writer.sheets[sheet_name], writer)
                    else:
                        logger.error(f"Не удалось сохранить таблицу '{sheet_name}' - неверный формат данных")
        except Exception as e:
            logger.error(f"Ошибка при сохранении таблиц в {excel_path}: {str(e)}")
    
    logger.debug(f"Таблицы сохранены в {excel_path}")