# -*- coding: utf-8 -*-
"""
Chg_XLS2MD_2309: Новый экспорт Excel -> MD в кастомном формате строк.
Формат MD:
1-4 строки: шапка (идентична прежней, дата обновляется)
Далее повторяющиеся блоки строк для каждой строки таблицы:
=== ROW ===
Field1:Value1
Field2:Value2
...

Правила экранирования при записи:
- \\  -> \\\\ (дублируем обратный слэш)
- |   -> \|   (экранируем вертикальную черту)
- :   -> \:   (экранируем двоеточие)
- \n  -> \\n (буквальная последовательность для перевода строки)

Все значения пишутся в одну строку. Пустые значения записываются как "Field:" (без значения).
"""

import os
from datetime import datetime
from openpyxl import load_workbook

# Константы путей по умолчанию
DOCS_DIR = r"c:\DEV\hh-applicant-tool\hh_v3\v4\docs"
XLSX_PATH = os.path.join(DOCS_DIR, "req.xlsx")
ROW_DELIM = "=== ROW ==="

# Chg_XLS2MD_2309: экранирование значений для MD
def _escape_md_value(value: str) -> str:
    if value is None:
        return ""
    s = str(value)
    s = s.replace("\\", "\\\\")  # escape backslash first
    s = s.replace("|", "\\|")
    s = s.replace(":", "\\:")
    s = s.replace("\n", "\\n")
    return s

# Chg_XLS2MD_2309: генерация шапки MD
def _generate_md_header(now: datetime) -> str:
    # Формируем первые 4 строки в стиле исходного файла
    dt_str = now.strftime("%d.%m.%Y %H:%M")
    header_lines = [
        "# Requirements Consolidated Table (Экспорт для Excel)",
        "",
        f"> Экспортировано из Excel req.xlsx (дата: {dt_str})",
        "",
    ]
    return "\n".join(header_lines)

# Chg_XLS2MD_2309: экспорт рабочей книги в MD
def xlsx_to_md(excel_path: str = XLSX_PATH, out_dir: str = DOCS_DIR) -> str:
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active

    # Заголовки из первой строки
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value is not None else "")

    now = datetime.now()
    ts_name = now.strftime("%H%M%d%m")  # HHMMDDMM
    out_name = f"req_{ts_name}.md"
    out_path = os.path.join(out_dir, out_name)

    lines = []
    lines.append(_generate_md_header(now))

    max_row = ws.max_row
    max_col = ws.max_column

    # Проходим по всем строкам начиная со второй
    for r in range(2, max_row + 1):
        # Считываем значения строки
        row_vals = []
        is_all_empty = True
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if val is None or str(val).strip() == "":
                row_vals.append("")
            else:
                is_all_empty = False
                row_vals.append(str(val))
        # Пропускаем полностью пустые строки (хвост)
        if is_all_empty:
            continue

        # Пишем разделитель строки
        lines.append(ROW_DELIM)
        # Пишем пары Field:Value
        for h, v in zip(headers, row_vals):
            # Поле может быть пустым в хедерах — пропускаем такие технические колонки
            field = (h or "").strip()
            if not field:
                continue
            lines.append(f"{field}:{_escape_md_value(v)}")

    # Сохраняем файл
    os.makedirs(out_dir, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    print(f"MD exported: {out_path}")
    return out_path

if __name__ == "__main__":
    # Простой сценарий запуска из IDE/консоли
    xlsx_to_md()
