import pandas as pd
import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl.styles
import numpy as np

# Chg_MDParser_2309: настройки нового формата MD
ROW_DELIM = "=== ROW ==="
DOCS_DIR = r"c:\DEV\hh-applicant-tool\hh_v3\v4\docs"
DEFAULT_OLD_MD = os.path.join(DOCS_DIR, "Requirements_Consolidated_Table.md")

# Chg_MDParser_2309: вспомогательная функция — найти последнюю req_*.md
def _find_latest_req_md(docs_dir):
    try:
        candidates = [
            os.path.join(docs_dir, f)
            for f in os.listdir(docs_dir)
            if f.startswith("req_") and f.endswith(".md")
        ]
        if not candidates:
            return None
        candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return candidates[0]
    except Exception:
        return None

# Chg_MDParser_2309: раскодировать экранирование из MD (обратное к xlsx_to_md)
def _unescape_md_value(s):
    if s is None:
        return ""
    # Сначала заменяем литералы перевода строки, затем спецсимволы, затем обратный слэш
    s = s.replace("\\n", "\n")
    s = s.replace("\\|", "|")
    s = s.replace("\\:", ":")
    s = s.replace("\\\\", "\\")
    return s

# Chg_MDParser_2309: найти индекс первого НЕэкранированного двоеточия
def _find_unescaped_colon(line):
    bs = 0
    for i, ch in enumerate(line):
        if ch == "\\":
            bs += 1
            continue
        if ch == ":" and (bs % 2 == 0):
            return i
        bs = 0
    return -1

# Chg_MDParser_2309: детект старого табличного формата MD
def _looks_like_old_table(md_path):
    try:
        with open(md_path, "r", encoding="utf-8") as f:
            lines = f.read().splitlines()
        # После 4 служебных строк должна быть строка с заголовками через | и затем |---|
        if len(lines) >= 6 and lines[4].strip().startswith("|") and "|---" in lines[5]:
            return True
        return False
    except Exception:
        return False

# Chg_MDParser_2309: парсер нового формата в список словарей
def _parse_new_md(md_path):
    with open(md_path, "r", encoding="utf-8") as f:
        lines = [ln.rstrip("\n") for ln in f]

    rows = []
    current = {}

    # Пропускаем первые 4 строки шапки
    i = 4 if len(lines) >= 4 else 0
    while i < len(lines):
        line = lines[i]
        if line.strip() == ROW_DELIM:
            # Начинается новая запись — сбрасываем предыдущую, если есть данные
            if current:
                rows.append(current)
                current = {}
            i += 1
            continue

        if not line.strip():
            # Пустая строка — абзац. В логике очистки это просто разделитель, пропускаем.
            i += 1
            continue

        # Ожидаем формат "Поле:Значение"
        idx = _find_unescaped_colon(line)
        if idx == -1:
            # Нет двоеточия — некорректная строка, пропускаем
            i += 1
            continue
        raw_field = line[:idx].strip()
        raw_value = line[idx + 1 :].lstrip()  # допускаем пробел после двоеточия
        field = _unescape_md_value(raw_field)
        value = _unescape_md_value(raw_value)

        # Правило очистки: если записано просто "Поле:" и далее абзац — значение пустое
        # В нашем представлении пустое значение уже пустое, т.к. raw_value == ""
        current[field] = value
        i += 1

    # Финальный пуш последней записи
    if current:
        rows.append(current)

    return rows

# Chg_MDParser_2309: получить канонические заголовки из старого MD (строка заголовка после 4 служебных)
def _get_canonical_headers_from_old_md(md_path):
    try:
        with open(md_path, "r", encoding="utf-8") as f:
            lines = f.read().splitlines()
        if len(lines) < 6:
            return []
        header_line = lines[4].strip()
        if not header_line.startswith("|"):
            return []
        # Разбиваем по | и чистим пробелы, отбрасываем пустые
        parts = [p.strip() for p in header_line.split("|")]
        headers = [p for p in parts if p]
        return headers
    except Exception:
        return []

# Прямой вызов для обновления существующего req.xlsx с сохранением форматирования
input_file = r"c:\DEV\hh-applicant-tool\hh_v3\v4\docs\req_16572309_final.md"
output_file = r"c:\DEV\hh-applicant-tool\hh_v3\v4\docs\req — копия.xlsx"

# Chg_MDParser_2309: если присутствуют файлы req_*.md — используем последний
_latest_req = _find_latest_req_md(DOCS_DIR)
if _latest_req:
    input_file = _latest_req

# Chg_MDParser_2309: определяем режим парсинга
use_old_table = _looks_like_old_table(input_file)

df_md = None
parsed_rows = None
if use_old_table:
    # Старый табличный формат через pandas
    df_md = pd.read_csv(input_file, sep='|', engine='python', skiprows=4, header=0, skipinitialspace=True)
    df_md.columns = df_md.columns.str.strip()
    # Удаляем пустые колонки, если есть
    df_md = df_md.dropna(axis=1, how='all')
    # Заменяем пустые строки на NaN для корректного присвоения
    df_md = df_md.replace('', np.nan)
else:
    # Новый формат Field:Value + разделители строк
    parsed_rows = _parse_new_md(input_file)

# Проверка, существует ли Excel-файл
if os.path.exists(output_file):
    # Загружаем существующий Excel с сохранением форматирования
    wb = load_workbook(output_file)
    ws = wb.active  # Предполагаем, что данные на первом листе

    print(f"Файл {output_file} существует. Обновляю с сохранением форматирования...")

    # Очищаем существующие данные, начиная со второй строки (сохраняем заголовки)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # Получаем заголовки из Excel (строгое имя столбцов)
    excel_headers = [cell.value for cell in ws[1]]
    normalized_headers = [(h or '').strip() for h in excel_headers]

    # Chg_MDParser_2309: подготовка итератора по строкам из MD
    md_rows_iter = []
    if use_old_table and df_md is not None:
        # Преобразуем DataFrame в список dict по заголовкам
        for _, row_md in df_md.iterrows():
            md_rows_iter.append({str(k): (row_md[k] if pd.notna(row_md[k]) else '') for k in df_md.columns})
    else:
        md_rows_iter = parsed_rows or []

    # Добавляем/обновляем строки из MD
    start_row = 2  # Начинаем со второй строки
    for row_md_map in md_rows_iter:
        # Нормализуем ключи полей
        norm_map = {(k or '').strip(): ('' if row_md_map[k] is None else str(row_md_map[k])) for k in row_md_map.keys()}
        req_id = norm_map.get('Requirement ID', '')

        # Ищем по Requirement ID в существующих строках
        found = False
        if req_id:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1):
                if row[0].value == req_id:
                    # Обновляем строку: теперь очищаем отсутствующие поля тоже
                    for col_idx, header in enumerate(excel_headers):
                        header_norm = normalized_headers[col_idx]
                        if not header_norm:
                            continue
                        value = norm_map.get(header_norm, '')
                        ws.cell(row=row[0].row, column=col_idx + 1, value=value)
                    found = True
                    print(f"Обновлена строка: {req_id}")
                    break

        if not found:
            # Добавляем новую строку, выставляя пустые значения для пропущенных полей
            for col_idx, header in enumerate(excel_headers):
                header_norm = normalized_headers[col_idx]
                if not header_norm:
                    continue
                value = norm_map.get(header_norm, '')
                ws.cell(row=start_row, column=col_idx + 1, value=value)
            print(f"Добавлена строка: {req_id}")
            start_row += 1

    # Сохраняем с сохранением форматирования
    wb.save(output_file)
    print(f"Обновление завершено. Файл сохранен как {output_file} с сохранением форматирования")
else:
    # Если файла нет, создаем новый
    if df_md is not None:
        # Старый табличный формат — можно напрямую в Excel
        df_md.to_excel(output_file, index=False, engine='openpyxl')
        print(f"Файл {output_file} не существует. Создаю новый из табличного MD...")
    else:
        # Новый формат — создаем книгу и заголовки
        headers = _get_canonical_headers_from_old_md(DEFAULT_OLD_MD)
        if not headers:
            # Строим порядок заголовков из данных (в порядке появления ключей)
            seen = []
            for row_md_map in (parsed_rows or []):
                for k in row_md_map.keys():
                    key = (k or '').strip()
                    if key and key not in seen:
                        seen.append(key)
            headers = seen
            # Ставим Requirement ID первым, если присутствует
            if 'Requirement ID' in headers:
                headers = ['Requirement ID'] + [h for h in headers if h != 'Requirement ID']

        wb = Workbook()
        ws = wb.active
        # Заголовки
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        # Данные
        start_row = 2
        for row_md_map in (parsed_rows or []):
            norm_map = {(k or '').strip(): ('' if row_md_map[k] is None else str(row_md_map[k])) for k in row_md_map.keys()}
            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=start_row, column=col_idx, value=norm_map.get(h, ''))
            start_row += 1

        wb.save(output_file)
        print(f"Файл {output_file} не существовал. Создал новый и заполнил из MD.")