"""
CLI для HH Tool v4 - синхронная архитектура
Простые команды без сложных зависимостей
"""

import click
import json
import time
import logging
import os
import sys
import subprocess
from pathlib import Path
from typing import Dict, Optional
from logging.handlers import RotatingFileHandler
import psutil
import requests

from core.task_dispatcher import TaskDispatcher
from core.task_database import TaskDatabase
from core.models import SystemMonitor
from plugins.fetcher_v4 import FilterManager, estimate_total_pages, VacancyFetcher

# // Chg_LOG_ROTATE_1509: Настройка ротации логов (100 МБ, 3 архива)
Path('logs').mkdir(exist_ok=True)
_handlers = [
    RotatingFileHandler('logs/app.log', maxBytes=100*1024*1024, backupCount=3, encoding='utf-8'),
    logging.StreamHandler()
]
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=_handlers
)

@click.group()
@click.version_option(version='4.0.0')
def cli():
    """HH Applicant Tool v4 - Синхронный диспетчер задач"""
    pass

@cli.command()
@click.option('--workers', '-w', default=3, help='Количество worker threads')
@click.option('--chunk-size', '-c', default=500, help='Размер chunk для больших задач')
@click.option('--daemon', '-d', is_flag=True, help='Запуск в daemon режиме')
def start(workers: int, chunk_size: int, daemon: bool):
    """Запуск диспетчера задач"""
    
    # Создаём необходимые папки
    Path('logs').mkdir(exist_ok=True)
    Path('data').mkdir(exist_ok=True)
    
    click.echo(f"Запуск HH Tool v4 Dispatcher...")
    click.echo(f"Workers: {workers}, Chunk size: {chunk_size}")
    
    try:
        dispatcher = TaskDispatcher(max_workers=workers, chunk_size=chunk_size)
        
        if daemon:
            click.echo("Daemon режим не реализован, запуск в foreground")
        
        dispatcher.start()
        
    except KeyboardInterrupt:
        click.echo("\nОстановка по Ctrl+C...")
    except Exception as e:
        click.echo(f"Ошибка запуска: {e}", err=True)
        raise click.Abort()

@cli.command()
@click.option('--filter-id', '-f', help='ID конкретного фильтра')
@click.option('--max-pages', '-p', type=int, help='Максимум страниц для загрузки')
@click.option('--chunk-size', '-c', default=500, help='Размер chunk')
@click.option('--schedule-at', type=int, help='Unix timestamp для отложенного запуска')
def load_vacancies(filter_id: Optional[str], max_pages: Optional[int], 
                  chunk_size: int, schedule_at: Optional[int]):
    """Добавить задачу загрузки вакансий"""
    
    db = TaskDatabase()
    filter_manager = FilterManager()
    
    # Определяем фильтры для загрузки
    if filter_id:
        filters = [filter_manager.get_filter_by_id(filter_id)]
        if not filters[0]:
            click.echo(f"Фильтр {filter_id} не найден", err=True)
            raise click.Abort()
    else:
        filters = filter_manager.get_active_filters()
        if not filters:
            click.echo("Активные фильтры не найдены", err=True)
            raise click.Abort()
    
    click.echo(f"Создание задач загрузки для {len(filters)} фильтров...")
    
    # Создание задач для каждого фильтра
    for filter_data in filters:
        try:
            # Оценка количества страниц если не указано
            if not max_pages:
                fetcher = VacancyFetcher()
                estimated_pages = estimate_total_pages(filter_data, fetcher)
                pages_to_load = min(estimated_pages, 200)  # Разумное ограничение
            else:
                pages_to_load = max_pages
            
            task_params = {
                'filter': filter_data,
                'max_pages': pages_to_load,
                'chunk_size': chunk_size
            }
            
            # Добавление задачи (подключение к существующему диспетчеру или создание в БД)
            try:
                # Пытаемся добавить через активный диспетчер
                dispatcher = TaskDispatcher()
                task_id = dispatcher.add_task(
                    task_type='load_vacancies',
                    params=task_params,
                    schedule_at=schedule_at,
                    timeout_sec=3600  # 1 час на загрузку
                )
            except:
                # Если диспетчер не запущен, создаём задачу в БД
                import uuid
                task_id = str(uuid.uuid4())
                db.create_task(
                    task_id=task_id,
                    task_type='load_vacancies',
                    params=task_params,
                    schedule_at=schedule_at,
                    timeout_sec=3600
                )
            
            filter_name = filter_data.get('name', filter_data.get('id', 'unknown'))
            click.echo(f"✓ Создана задача {task_id[:8]}... для фильтра '{filter_name}' ({pages_to_load} страниц)")
            
        except Exception as e:
            click.echo(f"✗ Ошибка создания задачи для фильтра {filter_data.get('id')}: {e}", err=True)
    
    if schedule_at:
        click.echo(f"Задачи запланированы на {time.ctime(schedule_at)}")
    else:
        click.echo("Задачи добавлены в очередь для немедленного выполнения")

@cli.command()
@click.option('--limit', '-l', default=20, help='Количество задач для показа')
@click.option('--status', '-s', help='Фильтр по статусу (pending/running/completed/failed)')
def tasks(limit: int, status: Optional[str]):
    """Показать список задач"""
    
    db = TaskDatabase()
    
    # Получение задач
    with db.get_connection() as conn:
        query = "SELECT * FROM tasks"
        params = []
        
        if status:
            query += " WHERE status = ?"
            params.append(status)
        
        query += " ORDER BY created_at DESC LIMIT ?"
        params.append(limit)
        
        cursor = conn.execute(query, params)
        tasks_data = [dict(row) for row in cursor.fetchall()]
    
    if not tasks_data:
        click.echo("Задачи не найдены")
        return
    
    # Форматирование вывода
    click.echo(f"\n{'ID':<12} {'Type':<15} {'Status':<10} {'Created':<19} {'Progress'}")
    click.echo("-" * 80)
    
    for task in tasks_data:
        task_id = task['id'][:8] + "..."
        task_type = task['type']
        task_status = task['status']
        # Исправление обработки времени
        if task['created_at']:
            try:
                # Если это unix timestamp
                if task['created_at'] > 1000000000:  # После 2001 года
                    created_at = time.ctime(task['created_at'])[:19]
                else:
                    # Если это julian day - конвертируем
                    unix_time = (task['created_at'] - 2440587.5) * 86400
                    created_at = time.ctime(unix_time)[:19]
            except (ValueError, OverflowError, OSError):
                created_at = 'Invalid time'
        else:
            created_at = 'Unknown'
        
        # Прогресс
        progress_info = ""
        if task['progress_json']:
            try:
                progress = json.loads(task['progress_json'])
                if 'chunk_progress' in progress:
                    progress_info = progress['chunk_progress']
                elif 'current_page' in progress:
                    progress_info = f"page {progress['current_page']}"
            except:
                pass
        
        click.echo(f"{task_id:<12} {task_type:<15} {task_status:<10} {created_at:<19} {progress_info}")
    
    click.echo(f"\nПоказано {len(tasks_data)} задач")

@cli.command()
@click.argument('task_id')
def task_info(task_id: str):
    """Подробная информация о задаче"""
    
    db = TaskDatabase()
    task = db.get_task(task_id)
    
    if not task:
        click.echo(f"Задача {task_id} не найдена", err=True)
        raise click.Abort()
    
    # Основная информация
    click.echo(f"\n=== Задача {task['id']} ===")
    click.echo(f"Тип: {task['type']}")
    click.echo(f"Статус: {task['status']}")
    click.echo(f"Создана: {time.ctime(task['created_at'] * 86400 + time.mktime(time.gmtime(0))) if task['created_at'] else 'Unknown'}")
    
    if task['started_at']:
        click.echo(f"Запущена: {time.ctime(task['started_at'] * 86400 + time.mktime(time.gmtime(0)))}")
    
    if task['finished_at']:
        click.echo(f"Завершена: {time.ctime(task['finished_at'] * 86400 + time.mktime(time.gmtime(0)))}")
    
    click.echo(f"Таймаут: {task['timeout_sec']} сек")
    
    # Параметры
    if task.get('params'):
        click.echo(f"\nПараметры:")
        for key, value in task['params'].items():
            if key == 'filter' and isinstance(value, dict):
                filter_name = value.get('name', value.get('id', 'unknown'))
                click.echo(f"  {key}: {filter_name}")
            else:
                click.echo(f"  {key}: {value}")

@cli.command()
@click.argument('output_path', type=click.Path())
@click.option('--format', '-f', default='brief', 
              type=click.Choice(['brief', 'full', 'analytical']),
              help='Формат экспорта: brief (краткий), full (полный), analytical (аналитический)')
@click.option('--limit', '-l', type=int, help='Максимальное количество записей')
@click.option('--date-from', type=str, help='Дата от (YYYY-MM-DD)')
@click.option('--min-salary', type=int, help='Минимальная зарплата')
@click.option('--area', type=str, help='Город/регион (частичное совпадение)')
@click.option('--include-description', is_flag=True, help='Включить описания вакансий (увеличивает размер файла)')
@click.option('--show-formats', is_flag=True, help='Показать доступные форматы экспорта')
def export(output_path: str, format: str, limit: Optional[int], date_from: Optional[str], 
          min_salary: Optional[int], area: Optional[str], include_description: bool, show_formats: bool):
    """Экспорт вакансий в Excel файл с оптимизацией размера"""
    
    # Показываем доступные форматы
    if show_formats:
        try:
            from core.export import VacancyExporter
            exporter = VacancyExporter()
            formats = exporter.get_export_formats()
            
            click.echo("\n📋 Доступные форматы экспорта:")
            for fmt_key, fmt_info in formats.items():
                click.echo(f"  {fmt_key:12} - {fmt_info['name']}")
                click.echo(f"             {fmt_info['description']}")
                click.echo(f"             Колонок: {len(fmt_info['columns'])}")
            click.echo()
            return
        except ImportError as e:
            click.echo(f"❌ Ошибка импорта экспортера: {e}", err=True)
            return
    
    try:
        from core.export import VacancyExporter
        
        # Создаем экспортер
        exporter = VacancyExporter()
        
        # Подготавливаем фильтры
        filters = {}
        if date_from:
            filters['date_from'] = date_from
        if min_salary:
            filters['min_salary'] = min_salary
        if area:
            filters['area_name'] = area
        
        # Проверяем количество записей
        total_count = exporter.get_vacancy_count(filters if filters else None)
        export_count = min(total_count, limit) if limit else total_count
        
        if total_count == 0:
            click.echo("❌ Нет вакансий для экспорта")
            return
        
        click.echo(f"📊 Найдено {total_count} вакансий в БД")
        if limit and limit < total_count:
            click.echo(f"   Будет экспортировано: {export_count} (ограничение)")
        else:
            click.echo(f"   Будет экспортировано: {export_count}")
        
        if filters:
            click.echo("🔍 Активные фильтры:")
            for key, value in filters.items():
                click.echo(f"   {key}: {value}")
        
        # Предупреждение о размере файла
        if export_count > 1000 and not limit:
            click.echo("⚠️  Большое количество записей может создать файл >50МБ")
            if not click.confirm("Продолжить экспорт?"):
                return
        
        click.echo(f"\n🚀 Начинаем экспорт в формате '{format}'...")
        
        # Выполняем экспорт
        result = exporter.export_to_excel(
            output_path=output_path,
            format_type=format,
            limit=limit,
            filters=filters if filters else None,
            include_description=include_description
        )
        
        # Выводим результаты
        if result['success']:
            click.echo(f"✅ Экспорт завершен успешно!")
            click.echo(f"   Файл: {result['file_path']}")
            click.echo(f"   Записей: {result['records_exported']}")
            click.echo(f"   Размер: {result['file_size_mb']} МБ")
            click.echo(f"   Время: {result['export_time_seconds']} сек")
            
            # Проверяем цель по размеру файла
            if result['file_size_mb'] > 50:
                click.echo(f"⚠️  Размер файла превышает цель 50МБ")
            else:
                click.echo(f"🎯 Размер файла соответствует цели (<50МБ)")
            
            # // Chg_EXPORT_VERIFY_2009: Верификация результата Excel
            try:
                import openpyxl
                from pathlib import Path
                from typing import Any
                
                xlsx_path = Path(result['file_path'])
                wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
                sheet = wb[wb.sheetnames[0]]
                
                headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
                data_rows = 0
                first_row = None
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not all((v is None or str(v).strip() == '') for v in row):
                        data_rows += 1
                        if first_row is None:
                            first_row = list(row)
                wb.close()
                
                click.echo("\n🔎 Проверка созданного файла:")
                click.echo(f"   Заголовки: {headers}")
                click.echo(f"   Строк данных (без заголовка): {data_rows}")
                if first_row is not None:
                    click.echo(f"   Первая строка: {first_row}")
                
                if data_rows < 10:
                    click.echo("❌ В файле меньше 10 строк данных — проверим фильтры/данные БД")
                else:
                    click.echo("✅ Данных достаточно (>=10 строк) — можно переходить к следующему шагу")
            except Exception as e:
                click.echo(f"⚠️  Не удалось автоматически проверить Excel: {e}")
                
        else:
            click.echo(f"❌ Ошибки при экспорте:")
            for error in result['errors']:
                click.echo(f"   • {error}")
    
    except ImportError as e:
        click.echo(f"❌ Модуль экспорта недоступен: {e}", err=True)
        click.echo("💡 Установите зависимости: pip install openpyxl pandas", err=True)
    except Exception as e:
        click.echo(f"❌ Ошибка экспорта: {e}", err=True)
        if logging.getLogger().isEnabledFor(logging.DEBUG):
            import traceback
            click.echo(traceback.format_exc(), err=True)

@cli.command()
@click.argument('test_type', type=click.Choice(['consolidated', 'diagnostic', 'legacy']), default='consolidated')
@click.option('--priority', default='1,2', help='Приоритеты тестов (1,2,3)')
@click.option('--output', type=str, help='Файл для сохранения JSON отчета')
@click.option('--verbose', '-v', is_flag=True, help='Подробный вывод')
def test(test_type: str, priority: str, output: Optional[str], verbose: bool):
    """Запуск консолидированных тестов v4"""
    
    if test_type == 'consolidated':
        click.echo("🚀 Запуск консолидированных тестов HH v4")
        
        try:
            # Импорт и запуск консолидированных тестов
            sys.path.insert(0, str(Path(__file__).parent))
            from tests.consolidated_tests import TestRunner
            
            # Парсинг приоритетов
            priorities = [int(p.strip()) for p in priority.split(',')]
            
            runner = TestRunner(priorities)
            results = runner.run_all_tests()
            
            # Сохранение в файл если указан
            if output:
                output_path = Path(output)
                output_path.parent.mkdir(exist_ok=True)
                
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(results, f, indent=2, ensure_ascii=False)
                click.echo(f"📋 Результаты сохранены в {output}")
            
            # Определение кода выхода на основе результатов
            if results['overall_percentage'] >= 90:
                click.echo(click.style("🎉 Тесты пройдены успешно!", fg='green'))
                return 0
            elif results['overall_percentage'] >= 70:
                click.echo(click.style("⚠️  Тесты пройдены с предупреждениями", fg='yellow'))
                return 0
            else:
                click.echo(click.style("❌ Критические проблемы в тестах", fg='red'))
                return 1
                
        except ImportError as e:
            click.echo(f"❌ Ошибка импорта тестов: {e}")
            return 1
        except Exception as e:
            click.echo(f"❌ Ошибка выполнения тестов: {e}")
            if verbose:
                import traceback
                click.echo(traceback.format_exc())
            return 1
    
    elif test_type == 'diagnostic':
        click.echo("🔍 Запуск системной диагностики HH v4")
        
        try:
            sys.path.insert(0, str(Path(__file__).parent))
            from tests.diagnostic_tests import SystemDiagnostic
            
            diagnostic = SystemDiagnostic()
            report = diagnostic.run_full_diagnostic()
            
            # Сохранение отчета
            if output:
                output_path = Path(output)
                output_path.parent.mkdir(exist_ok=True)
                
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(report, f, indent=2, ensure_ascii=False)
                click.echo(f"📋 Диагностический отчет сохранен в {output}")
            
            # Код выхода на основе здоровья системы
            if report['health_score'] >= 90:
                return 0
            elif report['health_score'] >= 70:
                return 2  # Предупреждения
            else:
                return 1  # Критические проблемы
                
        except ImportError as e:
            click.echo(f"❌ Ошибка импорта диагностики: {e}")
            return 1
        except Exception as e:
            click.echo(f"❌ Ошибка диагностики: {e}")
            if verbose:
                import traceback
                click.echo(traceback.format_exc())
            return 1
    
    elif test_type == 'legacy':
        # Старые простые тесты для совместимости
        click.echo("🔧 Запуск legacy тестов (простая проверка)")
        
        # Тест базы данных
        try:
            db = TaskDatabase()
            with db.get_connection():
                click.echo("✓ База данных доступна")
        except Exception as e:
            click.echo(f"✗ База данных: {e}")
        
        # Тест системных ресурсов
        try:
            monitor = SystemMonitor()
            metrics = monitor.get_system_metrics()
            cpu_usage = metrics.get('cpu_percent', 0)
            memory_usage = metrics.get('memory_percent', 0)
            
            click.echo(f"✓ Системные ресурсы: CPU {cpu_usage:.1f}%, RAM {memory_usage:.1f}%")
            
            if cpu_usage > 90 or memory_usage > 90:
                click.echo("⚠ Высокая нагрузка на систему")
                
        except Exception as e:
            click.echo(f"✗ Системные ресурсы: {e}")
        
        # Тест конфигурации
        try:
            config_path = Path('config/config_v4.json')
            if config_path.exists():
                with open(config_path) as f:
                    json.load(f)  # Валидация JSON
                click.echo("✓ Конфигурация загружена")
            else:
                click.echo("✗ Файл конфигурации не найден")
        except Exception as e:
            click.echo(f"✗ Конфигурация: {e}")
        
        click.echo("\n💡 Для полного тестирования используйте: python cli_v4.py test consolidated")

@cli.command()
@click.option('--suite', default='all', help='Набор тестов для запуска (all, readiness, unit)')
@click.option('--verbose', '-v', is_flag=True, help='Подробный вывод')
def test_suite(suite: str, verbose: bool):
    """Запуск автоматических тестов"""
    import subprocess
    import sys
    from pathlib import Path

    click.echo(f"🧪 Запуск тестов: {suite}")

    if suite == 'readiness':
        # Запуск тестов готовности системы
        test_file = Path(__file__).parent / "tests" / "test_system_readiness.py"

        if test_file.exists():
            try:
                # Прямой запуск скрипта
                result = subprocess.run(
                    [sys.executable, str(test_file)],
                    capture_output=True, text=True, cwd=Path(__file__).parent
                )

                click.echo(result.stdout)
                if result.stderr:
                    click.echo(click.style(result.stderr, fg='red'))

                if result.returncode == 0:
                    click.echo(click.style("✅ Тесты готовности прошли успешно!", fg='green'))
                else:
                    click.echo(click.style("❌ Некоторые тесты провалились", fg='red'))
                    sys.exit(1)

            except Exception as e:
                click.echo(click.style(f"❌ Ошибка запуска тестов: {e}", fg='red'))
                sys.exit(1)
        else:
            click.echo(click.style("❌ Файл тестов не найден", fg='red'))
            sys.exit(1)

    elif suite == 'all':
        # Попытка запуска через pytest
        try:
            cmd = [sys.executable, '-m', 'pytest', 'tests/', '-v' if verbose else '-q']
            result = subprocess.run(cmd, cwd=Path(__file__).parent)
            sys.exit(result.returncode)
        except FileNotFoundError:
            click.echo(click.style("⚠️  pytest не установлен, запускаем тесты готовности", fg='yellow'))
            # Fallback на тесты готовности
            ctx = click.get_current_context()
            ctx.invoke(test_suite, suite='readiness', verbose=verbose)

    else:
        click.echo(click.style(f"❌ Неизвестный набор тестов: {suite}", fg='red'))


@cli.command()
@click.option('--type', 'cleanup_type', default='files', 
              type=click.Choice(['files', 'logs', 'archives', 'all']),
              help='Тип очистки')
@click.option('--days', default=14, help='Удалить файлы старше N дней')
@click.option('--dry-run', is_flag=True, help='Показать что будет удалено, не удаляя')
def cleanup(cleanup_type: str, days: int, dry_run: bool):
    """Очистка временных файлов и устаревших данных"""
    from pathlib import Path
    import time
    import shutil
    
    click.echo(f"🧹 Очистка: {cleanup_type} (старше {days} дней)")
    if dry_run:
        click.echo("📋 РЕЖИМ ПРЕДВАРИТЕЛЬНОГО ПРОСМОТРА - файлы не будут удалены")
    
    base_path = Path(__file__).parent
    quarantine_dir = base_path / "data" / ".trash"
    
    if not dry_run:
        quarantine_dir.mkdir(parents=True, exist_ok=True)
    
    cleanup_stats = {"moved": 0, "deleted": 0, "errors": []}
    cutoff_time = time.time() - (days * 24 * 60 * 60)
    
    def should_cleanup(file_path: Path) -> bool:
        """Проверить, нужно ли удалять файл"""
        try:
            return file_path.stat().st_mtime < cutoff_time
        except:
            return False
    
    def safe_move_to_quarantine(file_path: Path):
        """Безопасное перемещение в карантин"""
        try:
            if dry_run:
                click.echo(f"  🗑️  {file_path}")
                cleanup_stats["moved"] += 1
            else:
                quarantine_path = quarantine_dir / file_path.name
                # Избегаем конфликтов имен
                counter = 1
                while quarantine_path.exists():
                    name = f"{file_path.stem}_{counter}{file_path.suffix}"
                    quarantine_path = quarantine_dir / name
                    counter += 1
                
                shutil.move(str(file_path), str(quarantine_path))
                cleanup_stats["moved"] += 1
                click.echo(f"  📦 {file_path} → карантин")
        except Exception as e:
            cleanup_stats["errors"].append(f"{file_path}: {e}")
    
    # Очистка временных файлов
    if cleanup_type in ['files', 'all']:
        click.echo("\n📁 Поиск временных файлов...")
        
        # Временные файлы в корне и data/
        for pattern in ['*.tmp', '*.bak']:
            for file_path in base_path.glob(pattern):
                if should_cleanup(file_path):
                    safe_move_to_quarantine(file_path)
    
    # Очистка логов
    if cleanup_type in ['logs', 'all']:
        click.echo("\n📋 Поиск старых логов...")
        logs_dir = base_path / "logs"
        if logs_dir.exists():
            for log_file in logs_dir.glob("*.log"):
                if should_cleanup(log_file):
                    safe_move_to_quarantine(log_file)
    
    # Отчет
    click.echo(f"\n📊 Результаты очистки:")
    click.echo(f"  Перемещено в карантин: {cleanup_stats['moved']}")
    if cleanup_stats['errors']:
        click.echo(f"  Ошибки: {len(cleanup_stats['errors'])}")


@cli.command()
def status():
    """Показать общий статус системы"""
    
    db = TaskDatabase()
    stats = db.get_stats()
    
    click.echo("\n=== Статус HH Tool v4 ===")
    
    # Статистика задач
    click.echo("\nЗадачи за последний день:")
    if stats.get('tasks'):
        for status, count in stats['tasks'].items():
            click.echo(f"  {status}: {count}")
    else:
        click.echo("  Нет задач")
    
    # Статистика вакансий
    click.echo("\nВакансии:")
    vacancy_stats = db.get_vacancy_stats()
    click.echo(f"  Всего: {vacancy_stats.get('total_vacancies', 0)}")
    click.echo(f"  Обработано: {vacancy_stats.get('processed_vacancies', 0)}")
    click.echo(f"  Сегодня загружено: {vacancy_stats.get('today_vacancies', 0)}")
    
    # Статистика по фильтрам
    filter_stats = db.get_vacancy_count_by_filter()
    if filter_stats:
        click.echo("\nВакансии по фильтрам (последние 7 дней):")
        for filter_id, count in list(filter_stats.items())[:10]:  # Топ 10
            click.echo(f"  {filter_id}: {count}")
    
    click.echo(f"\nОбновлено: {stats.get('timestamp', 'Unknown')}")


@cli.command()
@click.option('--days', '-d', default=7, help='Количество дней для статистики (по умолчанию: 7)')
@click.option('--format', '-f', 'output_format', default='table', 
              type=click.Choice(['table', 'json']), help='Формат вывода')
@click.option('--changes-only', '-c', is_flag=True, help='Показать только статистику изменений')
def stats(days: int, output_format: str, changes_only: bool):
    """Статистика версионирования и изменений данных"""
    
    try:
        from core.task_database import TaskDatabase
        db = TaskDatabase()
        changes_stats = db.get_combined_changes_stats(days)
        
        if output_format == 'json':
            import json
            click.echo(json.dumps(changes_stats, ensure_ascii=False, indent=2))
            return
        
        # Форматированный вывод
        click.echo(f"\n📊 === СТАТИСТИКА ИЗМЕНЕНИЙ ЗА {days} ДНЕЙ (v4) ===")
        
        # Вакансии
        vacancy_stats = changes_stats.get('vacancies', {})
        click.echo(f"\n🔍 Вакансии:")
        click.echo(f"  ✅ Новых вакансий: {vacancy_stats.get('new_vacancies', 0)}")
        click.echo(f"  🔄 Новых версий: {vacancy_stats.get('new_versions', 0)}")
        click.echo(f"  ⏭️  Дубликатов пропущено: {vacancy_stats.get('duplicates_skipped', 0)}")
        click.echo(f"  📈 Эффективность: {vacancy_stats.get('efficiency_percentage', 0)}%")
        click.echo(f"  📊 Всего операций: {vacancy_stats.get('total_changes', 0)}")
        
        # Работодатели
        employer_stats = changes_stats.get('employers', {})
        if employer_stats.get('total_changes', 0) > 0:
            click.echo(f"\n🏢 Работодатели:")
            click.echo(f"  📊 Всего операций: {employer_stats.get('total_changes', 0)}")
        
        # Сводка
        summary = changes_stats.get('summary', {})
        click.echo(f"\n🎯 Итого:")
        click.echo(f"  📋 Всего операций: {summary.get('total_operations', 0)}")
        
        if not changes_only:
            # Общая статистика БД
            click.echo(f"\n💾 База данных:")
            try:
                db_stats = db.get_stats()
                click.echo(f"  📦 Всего вакансий: {db_stats.get('total_vacancies', 0)}")
                click.echo(f"  🗄️  Размер БД: {db_stats.get('db_size_mb', 0)} МБ")
            except Exception:
                pass
        
        # Показать детали при малом количестве изменений
        if vacancy_stats.get('total_changes', 0) < 10:
            click.echo(f"\n⚠️  Мало изменений за {days} дней. Попробуйте увеличить период или проверить загрузку данных.")
        
    except ImportError as e:
        click.echo(f"❌ Ошибка импорта: {e}", err=True)
    except Exception as e:
        click.echo(f"❌ Ошибка получения статистики: {e}", err=True)
        if click.get_current_context().obj and click.get_current_context().obj.get('debug'):
            import traceback
            click.echo(traceback.format_exc(), err=True)


@cli.command()
@click.option('--detailed', '-d', is_flag=True, help='Детальная информация о системе')
@click.option('--alerts-only', '-a', is_flag=True, help='Показать только алерты')
@click.option('--json-format', '-j', is_flag=True, help='Вывод в JSON формате')
def system(detailed: bool, alerts_only: bool, json_format: bool):
    """Системный мониторинг и диагностика"""
    
    try:
        monitor = SystemMonitor()
        
        if alerts_only:
            # Показать только алерты
            metrics = monitor.get_comprehensive_metrics()
            alerts = metrics.get('alerts', [])
            
            if json_format:
                click.echo(json.dumps({'alerts': alerts}, ensure_ascii=False, indent=2))
            else:
                if alerts:
                    click.echo(f"\n🚨 Активные алерты ({len(alerts)}):")
                    for alert in alerts:
                        level_icon = {'info': 'ℹ️', 'warning': '⚠️', 'critical': '🔥'}.get(alert['level'], '❓')
                        click.echo(f"  {level_icon} {alert['component']}: {alert['message']}")
                else:
                    click.echo("✅ Нет активных алертов")
            return
        
        if detailed:
            # Полная системная информация
            metrics = monitor.get_comprehensive_metrics()
            
            if json_format:
                click.echo(json.dumps(metrics, ensure_ascii=False, indent=2))
                return
            
            # Форматированный вывод
            click.echo("\n🖥️  === СИСТЕМНЫЙ МОНИТОРИНГ HH TOOL v4 ===")
            
            # Общий статус
            quick = monitor.get_quick_status()
            status_icon = {'healthy': '✅', 'warning': '⚠️', 'critical': '🔥', 'error': '❌'}.get(quick['overall_status'], '❓')
            click.echo(f"\n{status_icon} Общий статус: {quick['overall_status'].upper()}")
            click.echo(f"   CPU: {quick['cpu_percent']}% | Память: {quick['memory_percent']}%")
            
            # CPU информация
            system_data = metrics.get('system', {})
            cpu = system_data.get('cpu', {})
            if cpu and 'error' not in cpu:
                click.echo(f"\n💻 CPU:")
                click.echo(f"   Загрузка: {cpu['percent_total']}% ({cpu['count_logical']} логических ядер)")
                if cpu.get('load_average'):
                    la = cpu['load_average']
                    click.echo(f"   Load Average: {la['1min']}, {la['5min']}, {la['15min']}")
            
            # Память
            memory = system_data.get('memory', {})
            if memory and 'error' not in memory:
                virtual = memory.get('virtual', {})
                click.echo(f"\n🧠 Память:")
                click.echo(f"   Виртуальная: {virtual.get('percent', 0)}% из {virtual.get('total_mb', 0)} МБ")
                click.echo(f"   Доступно: {virtual.get('available_mb', 0)} МБ")
            
            # База данных
            application = metrics.get('application', {})
            database = application.get('database', {})
            if database and database.get('status') == 'connected':
                click.echo(f"\n🗄️  База данных:")
                click.echo(f"   Размер: {database.get('file_size_mb', 0)} МБ")
                click.echo(f"   Режим: {database.get('journal_mode', 'unknown')}")
                tables = database.get('tables', {})
                total_records = sum(t.get('record_count', 0) for t in tables.values())
                click.echo(f"   Записей: {total_records} в {len(tables)} таблицах")
            
            # Health checks
            health_checks = application.get('health_checks', {})
            click.echo(f"\n🏥 Проверки здоровья:")
            for check_name, check_result in health_checks.items():
                status = check_result.get('status', 'unknown')
                message = check_result.get('message', 'No message')
                icon = {'pass': '✅', 'warning': '⚠️', 'fail': '❌'}.get(status, '❓')
                click.echo(f"   {icon} {check_name}: {message}")
            
            # Алерты
            alerts = metrics.get('alerts', [])
            if alerts:
                click.echo(f"\n🚨 Активные алерты ({len(alerts)}):")
                for alert in alerts:
                    level_icon = {'info': 'ℹ️', 'warning': '⚠️', 'critical': '🔥'}.get(alert['level'], '❓')
                    click.echo(f"   {level_icon} {alert['component']}: {alert['message']}")
        
        else:
            # Краткая информация (по умолчанию)
            quick = monitor.get_quick_status()
            
            if json_format:
                click.echo(json.dumps(quick, ensure_ascii=False, indent=2))
                return
            
            status_icon = {'healthy': '✅', 'warning': '⚠️', 'critical': '🔥', 'error': '❌'}.get(quick['overall_status'], '❓')
            click.echo(f"\n{status_icon} Статус системы: {quick['overall_status'].upper()}")
            click.echo(f"CPU: {quick['cpu_percent']}% | Память: {quick['memory_percent']}%")
            
            # Проверим наличие алертов
            metrics = monitor.get_comprehensive_metrics()
            alerts = metrics.get('alerts', [])
            if alerts:
                critical_count = len([a for a in alerts if a['level'] == 'critical'])
                warning_count = len([a for a in alerts if a['level'] == 'warning'])
                if critical_count:
                    click.echo(f"🔥 Критических алертов: {critical_count}")
                if warning_count:
                    click.echo(f"⚠️  Предупреждений: {warning_count}")
                click.echo(f"   Используйте --detailed для подробностей")
        
    except Exception as e:
        if json_format:
            click.echo(json.dumps({'error': str(e)}, ensure_ascii=False))
        else:
            click.echo(f"❌ Ошибка системного мониторинга: {e}", err=True)


@cli.command()
def filters():
    """Показать список фильтров"""
    
    filter_manager = FilterManager()
    filters_list = filter_manager.load_filters()
    
    if not filters_list:
        click.echo("Фильтры не найдены")
        return
    
    click.echo(f"\n{'ID':<15} {'Name':<30} {'Enabled':<8} {'Text'}")
    click.echo("-" * 80)
    
    for f in filters_list:
        filter_id = f.get('id', 'unknown')[:14]
        name = f.get('name', 'Unknown')[:29]
        enabled = "✓" if f.get('enabled', True) else "✗"
        text = f.get('text', '')[:30]
        
        click.echo(f"{filter_id:<15} {name:<30} {enabled:<8} {text}")
    
    click.echo(f"\nВсего фильтров: {len(filters_list)}")
    active_count = len([f for f in filters_list if f.get('enabled', True)])
    click.echo(f"Активных: {active_count}")

@cli.command()
@click.option('--host', default='localhost', help='Host для веб-интерфейса')
@click.option('--port', default=8080, help='Port для веб-интерфейса')
@click.option('--debug', is_flag=True, help='Режим отладки с автоперезагрузкой')
def dashboard(host: str, port: int, debug: bool):
    """Запуск улучшенной FastAPI веб-панели (как в v3)"""
    
    try:
        from web.server import run_web_server
        click.echo(f"🚀 Запуск HH Tool v4 Dashboard на http://{host}:{port}")
        click.echo("📊 Функции: WebSocket обновления, графики, детальная статистика")
        click.echo("⏹️  Для остановки нажмите Ctrl+C")
        
        run_web_server(host=host, port=port, debug=debug)
        
    except ImportError as e:
        click.echo(f"❌ Ошибка импорта: {e}", err=True)
        click.echo("💡 Установите зависимости: pip install fastapi uvicorn jinja2 websockets", err=True)
    except Exception as e:
        click.echo(f"❌ Ошибка запуска dashboard: {e}", err=True)

@cli.command()
@click.option('--host', default='localhost', help='Host для веб-интерфейса')
@click.option('--port', default=8000, help='Port для веб-интерфейса')
def web(host: str, port: int):
    """Запуск простого веб-интерфейса для мониторинга (legacy)"""
    
    try:
        from http.server import HTTPServer, BaseHTTPRequestHandler
        import urllib.parse
        
        class SimpleHandler(BaseHTTPRequestHandler):
            def do_GET(self):
                if self.path == '/':
                    self.send_response(200)
                    self.send_header('Content-type', 'text/html; charset=utf-8')
                    self.end_headers()
                    
                    # Простая HTML страница со статистикой
                    db = TaskDatabase()
                    stats = db.get_stats()
                    
                    html = f"""
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <title>HH Tool v4 Status</title>
                        <meta charset="utf-8">
                        <meta http-equiv="refresh" content="30">
                        <style>
                            body {{ font-family: Arial, sans-serif; margin: 40px; }}
                            .stats {{ background: #f5f5f5; padding: 20px; border-radius: 8px; margin: 20px 0; }}
                            .error {{ color: red; }}
                            .success {{ color: green; }}
                        </style>
                    </head>
                    <body>
                        <h1>HH Tool v4 - Статус системы</h1>
                        
                        <div class="stats">
                            <h2>Задачи за последний день</h2>
                            {self._format_tasks_stats(stats.get('tasks', {}))}
                        </div>
                        
                        <div class="stats">
                            <h2>Вакансии</h2>
                            {self._format_vacancy_stats(stats.get('vacancies', {}))}
                        </div>
                        
                        <p><small>Обновлено: {stats.get('timestamp', 'Unknown')} | Автообновление каждые 30 сек</small></p>
                    </body>
                    </html>
                    """
                    
                    self.wfile.write(html.encode('utf-8'))
                
                elif self.path == '/api/stats':
                    self.send_response(200)
                    self.send_header('Content-type', 'application/json')
                    self.end_headers()
                    
                    db = TaskDatabase()
                    stats = db.get_stats()
                    
                    self.wfile.write(json.dumps(stats, ensure_ascii=False).encode('utf-8'))
                
                else:
                    self.send_response(404)
                    self.end_headers()
            
            def _format_tasks_stats(self, tasks_stats):
                if not tasks_stats:
                    return "<p>Нет задач</p>"
                
                html = "<ul>"
                for status, count in tasks_stats.items():
                    css_class = "success" if status == "completed" else "error" if status == "failed" else ""
                    html += f'<li class="{css_class}">{status}: {count}</li>'
                html += "</ul>"
                return html
            
            def _format_vacancy_stats(self, vacancy_stats):
                html = "<ul>"
                html += f"<li>Всего: {vacancy_stats.get('total_vacancies', 0)}</li>"
                html += f"<li>Обработано: {vacancy_stats.get('processed_vacancies', 0)}</li>"
                html += f"<li>Сегодня загружено: {vacancy_stats.get('today_vacancies', 0)}</li>"
                html += "</ul>"
                return html
            
            def log_message(self, format, *args):
                pass  # Отключаем логи запросов
        
        server = HTTPServer((host, port), SimpleHandler)
        click.echo(f"Веб-интерфейс запущен на http://{host}:{port}")
        click.echo("Для остановки нажмите Ctrl+C")
        
        server.serve_forever()
        
    except ImportError:
        click.echo("Веб-интерфейс недоступен", err=True)
    except KeyboardInterrupt:
        click.echo("\nВеб-сервер остановлен")
    except Exception as e:
        click.echo(f"Ошибка запуска веб-сервера: {e}", err=True)

# // Chg_DEVUP_1509: Короткая команда dev-up для перезапуска панели и диспетчера
@cli.command(name='dev-up')
@click.option('--workers', '-w', default=2, help='Количество worker threads')
@click.option('--max-pages', '-p', default=1, help='Сколько страниц загрузить однократно')
@click.option('--no-load', is_flag=True, default=False, help='Не запускать разовую загрузку')
def dev_up(workers: int, max_pages: int, no_load: bool):
    """Убить процессы на 8080 и cli_v4 dashboard/start, поднять панель и диспетчер, опционально загрузить вакансии и показать статистику"""
    try:
        Path('logs').mkdir(exist_ok=True)
        Path('data').mkdir(exist_ok=True)

        # 1) Убиваем слушателей 8080
        killed = []
        try:
            for c in psutil.net_connections(kind='inet'):
                try:
                    if c.laddr and getattr(c.laddr, 'port', None) == 8080 and c.status == psutil.CONN_LISTEN and c.pid:
                        p = psutil.Process(c.pid)
                        p.kill()
                        killed.append(c.pid)
                except Exception:
                    pass
        except Exception:
            pass

        # 2) Убиваем старые процессы dashboard/start
        self_pid = os.getpid()
        for p in psutil.process_iter(['pid', 'name', 'cmdline']):
            try:
                if p.info['pid'] == self_pid:
                    continue
                cmd = ' '.join(p.info.get('cmdline') or [])
                if 'cli_v4.py' in cmd and ('dashboard' in cmd or 'start' in cmd):
                    p.kill()
                    killed.append(p.info['pid'])
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                continue

        click.echo(f"Убито процессов: {len(killed)}")

        # 3) Стартуем панель и диспетчер
        dash = subprocess.Popen([sys.executable, 'cli_v4.py', 'dashboard', '--host', 'localhost', '--port', '8080'], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        time.sleep(2)
        disp = subprocess.Popen([sys.executable, 'cli_v4.py', 'start', '--workers', str(workers)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        click.echo(f"Dashboard PID: {dash.pid}, Dispatcher PID: {disp.pid}")

        # 4) Однократная загрузка
        if not no_load:
            click.echo(f"Запуск однократной загрузки: {max_pages} стр.")
            subprocess.run([sys.executable, 'cli_v4.py', 'load-vacancies', '--max-pages', str(max_pages)], check=False)

        # 5) Ожидаем и выводим статистику
        ok = False
        for _ in range(12):
            try:
                r = requests.get('http://localhost:8080/api/stats', timeout=5)
                if r.ok:
                    data = r.json()
                    vac = data.get('vacancies', {})
                    click.echo(json.dumps({
                        'total_vacancies': vac.get('total_vacancies', 0),
                        'added_last_run_10m_window': vac.get('added_last_run_10m_window', 0),
                        'last_run_at': vac.get('last_run_at')
                    }, ensure_ascii=False))
                    ok = True
                    break
            except Exception:
                pass
            time.sleep(5)
    except Exception as e:
        click.echo(f"❌ Ошибка выполнения dev-up: {e}", err=True)

@cli.command()
@click.option('--host', '-h', help='Конкретный хост (host2, host3) или все')
@click.option('--enable', is_flag=True, help='Включить хост')
@click.option('--disable', is_flag=True, help='Выключить хост')
@click.option('--test', is_flag=True, help='Тестировать подключение')
@click.option('--status', is_flag=True, help='Показать статус')
def hosts(host: str, enable: bool, disable: bool, test: bool, status: bool):
    """Управление внешними хостами (Host2, Host3)"""
    import json
    from core.task_dispatcher import TaskDispatcher
    
    # Загружаем конфигурацию
    try:
        with open('config/config_v4.json', 'r', encoding='utf-8') as f:
            config = json.load(f)
    except FileNotFoundError:
        click.echo("❌ Файл конфигурации config/config_v4.json не найден")
        return
    except json.JSONDecodeError as e:
        click.echo(f"❌ Ошибка чтения конфигурации: {e}")
        return
    
    hosts_config = config.get('hosts', {})
    
    if not host:
        # Показываем статус всех хостов
        click.echo("🏠 === СТАТУС ХОСТОВ ===")
        click.echo()
        
        for host_id, host_config in hosts_config.items():
            name = host_config.get('name', host_id)
            description = host_config.get('description', 'Описание отсутствует')
            enabled = host_config.get('enabled', False)
            host_type = host_config.get('type', 'unknown')
            mock_mode = host_config.get('mock_mode', True)
            
            status_icon = "✅" if enabled else "❌"
            mock_text = " (MOCK)" if mock_mode else ""
            
            click.echo(f"{status_icon} {host_id.upper()}: {name}")
            click.echo(f"   📝 {description}")
            click.echo(f"   🔧 Тип: {host_type}{mock_text}")
            click.echo(f"   ⚡ Статус: {'Включен' if enabled else 'Выключен'}")
            click.echo()
        
        # Если включено тестирование, проверяем подключения
        if test:
            click.echo("🧪 === ТЕСТИРОВАНИЕ ПОДКЛЮЧЕНИЙ ===")
            dispatcher = TaskDispatcher(config=config)
            host_status = dispatcher.get_host_status()
            
            for host_id, status_info in host_status.items():
                status = status_info.get('status', 'unknown')
                host_type = status_info.get('type', 'unknown')
                
                if status == 'active':
                    click.echo(f"✅ {host_id.upper()}: Активен ({host_type})")
                elif status == 'healthy':
                    click.echo(f"✅ {host_id.upper()}: Здоров ({host_type})")
                elif status == 'disabled':
                    click.echo(f"⚠️  {host_id.upper()}: Отключен ({host_type})")
                else:
                    error_msg = status_info.get('error', 'Неизвестная ошибка')
                    click.echo(f"❌ {host_id.upper()}: Ошибка - {error_msg}")
        
        return
    
    # Операции с конкретным хостом
    if host not in hosts_config:
        click.echo(f"❌ Хост '{host}' не найден в конфигурации")
        available_hosts = ', '.join(hosts_config.keys())
        click.echo(f"💡 Доступные хосты: {available_hosts}")
        return
    
    host_config = hosts_config[host]
    host_name = host_config.get('name', host)
    
    if enable:
        hosts_config[host]['enabled'] = True
        click.echo(f"✅ Хост {host_name} включен")
        
        # Сохраняем конфигурацию
        try:
            with open('config/config_v4.json', 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
            click.echo("💾 Конфигурация сохранена")
        except Exception as e:
            click.echo(f"❌ Ошибка сохранения конфигурации: {e}")
    
    elif disable:
        hosts_config[host]['enabled'] = False
        click.echo(f"❌ Хост {host_name} выключен")
        
        # Сохраняем конфигурацию
        try:
            with open('config/config_v4.json', 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
            click.echo("💾 Конфигурация сохранена")
        except Exception as e:
            click.echo(f"❌ Ошибка сохранения конфигурации: {e}")
    
    elif test:
        click.echo(f"🧪 Тестирование {host_name}...")
        dispatcher = TaskDispatcher(config=config)
        
        if host == 'host2' and dispatcher.host2_client:
            try:
                health = dispatcher.host2_client.health_check()
                if health['status'] == 'healthy':
                    click.echo(f"✅ {host_name}: Подключение успешно")
                    click.echo(f"   📊 Режим: {'Mock' if health.get('mock_mode') else 'Real'}")
                    click.echo(f"   🔗 Адрес: {health.get('host')}:{health.get('port')}")
                else:
                    click.echo(f"❌ {host_name}: Проблемы с подключением")
            except Exception as e:
                click.echo(f"❌ {host_name}: Ошибка тестирования - {e}")
        
        elif host == 'host3' and dispatcher.host3_client:
            try:
                health = dispatcher.host3_client.health_check()
                if health['status'] == 'healthy':
                    click.echo(f"✅ {host_name}: Подключение успешно")
                    click.echo(f"   📊 Режим: {'Mock' if health.get('mock_mode') else 'Real'}")
                    click.echo(f"   🔗 Endpoint: {health.get('endpoint')}")
                    click.echo(f"   🤖 Модель: {health.get('model')}")
                else:
                    click.echo(f"❌ {host_name}: Сервис недоступен")
            except Exception as e:
                click.echo(f"❌ {host_name}: Ошибка тестирования - {e}")
        
        else:
            click.echo(f"⚠️  {host_name}: Хост отключен или не поддерживает тестирование")
    
    else:
        # Показываем информацию о конкретном хосте
        click.echo(f"🏠 === ИНФОРМАЦИЯ О ХОСТЕ {host.upper()} ===")
        click.echo(f"📝 Название: {host_config.get('name', host)}")
        click.echo(f"📋 Описание: {host_config.get('description', 'Не указано')}")
        click.echo(f"🔧 Тип: {host_config.get('type', 'unknown')}")
        click.echo(f"⚡ Включен: {'Да' if host_config.get('enabled') else 'Нет'}")
        click.echo(f"🎭 Mock режим: {'Да' if host_config.get('mock_mode') else 'Нет'}")
        
        if 'connection' in host_config:
            click.echo("🔗 Настройки подключения:")
            for key, value in host_config['connection'].items():
                if 'password' in key.lower() or 'key' in key.lower():
                    value = '***'
                click.echo(f"   {key}: {value}")


@cli.command()
@click.argument('action', type=click.Choice(['start', 'stop', 'status', 'restart']))
@click.option('--config', default='config/config_v4.json', help='Путь к конфигурации')
@click.option('--log-level', type=click.Choice(['DEBUG', 'INFO', 'WARNING', 'ERROR']), default='INFO', help='Уровень логирования')
@click.option('--background', is_flag=True, help='Запуск в фоновом режиме')
def daemon(action: str, config: str, log_level: str, background: bool):
    """Управление демоном планировщика"""
    import json
    import psutil
    import subprocess
    import signal
    from pathlib import Path
    from datetime import datetime
    
    pid_file = Path('data/scheduler_daemon.pid')
    
    if action == 'start':
        # Проверяем что демон не запущен
        if pid_file.exists():
            try:
                pid = int(pid_file.read_text().strip())
                if psutil.pid_exists(pid):
                    click.echo(f"⚠️  Найден работающий демон (PID: {pid}), останавливаем...")
                    # Принудительно останавливаем предыдущий процесс
                    try:
                        os.kill(pid, signal.SIGTERM)
                        import time
                        time.sleep(2)
                        if psutil.pid_exists(pid):
                            os.kill(pid, signal.SIGKILL)
                            time.sleep(1)
                        click.echo("✅ Предыдущий демон остановлен")
                    except:
                        pass
                    pid_file.unlink()
                else:
                    pid_file.unlink()  # Удаляем устаревший PID файл
            except:
                pid_file.unlink()
        
        # Очищаем зависшие процессы через БД
        click.echo("🔍 Очистка зависших процессов через БД...")
        try:
            from core.task_database import TaskDatabase
            db = TaskDatabase()
            db.cleanup_dead_processes()
            
            # Убиваем записанные процессы если они есть
            if db.kill_process("scheduler_daemon"):
                click.echo("🔪 Остановлен предыдущий демон из БД")
            if db.kill_process("web_server"):
                click.echo("🔪 Остановлен предыдущий веб-сервер из БД")
                
        except Exception as e:
            click.echo(f"⚠️  Ошибка очистки процессов: {e}")
        
        click.echo("🚀 Запуск демона планировщика...")
        
        if background:
            # Запуск в фоновом режиме
            cmd = [
                sys.executable, '-c',
                f'import sys; sys.path.insert(0, "."); '
                f'from core.scheduler_daemon import main; main()'
            ]
            
            # // Chg_UNIFIED_LOG_2009: Демон пишет в общий app.log
            try:
                process = subprocess.Popen(
                    cmd,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    cwd=Path.cwd(),
                    start_new_session=True
                )
                
                # // Chg_CLI_DAEMON_2009: Проверка реального запуска процесса
                import time
                time.sleep(1)  # Даём время процессу стартануть
                
                if process.poll() is None:  # Процесс ещё работает
                    # Дополнительная проверка через psutil
                    if psutil.pid_exists(process.pid):
                        pid_file.write_text(str(process.pid))
                        click.echo(f"✅ Демон запущен в фоновом режиме (PID: {process.pid})")
                        click.echo(f"📄 Логи: logs/app.log")
                    else:
                        click.echo(f"❌ Процесс демона не найден после запуска")
                        return
                else:
                    # Процесс завершился с ошибкой
                    return_code = process.poll()
                    click.echo(f"❌ Демон завершился с ошибкой (код: {return_code})")
                    
                    # Показываем последние строки общего лога
                    time.sleep(0.5)  # Даём время записать лог
                    try:
                        app_log = Path('logs/app.log')
                        if app_log.exists():
                            lines = app_log.read_text(encoding='utf-8').strip().split('\n')
                            click.echo("🔍 Последние записи лога:")
                            for line in lines[-5:]:
                                if line.strip():
                                    click.echo(f"   {line}")
                    except Exception:
                        pass
                    return
                    
            except Exception as e:
                click.echo(f"❌ Ошибка запуска процесса демона: {e}")
                return
            
        else:
            # Прямой запуск
            try:
                from core.scheduler_daemon import main
                main()
            except KeyboardInterrupt:
                click.echo("\n⏹️  Демон остановлен")
            except ImportError as e:
                click.echo(f"❌ Ошибка импорта: {e}")
            except Exception as e:
                click.echo(f"❌ Ошибка запуска демона: {e}")
    
    elif action == 'stop':
        if not pid_file.exists():
            click.echo("❌ Демон не запущен")
            return
        
        try:
            pid = int(pid_file.read_text().strip())
            
            if psutil.pid_exists(pid):
                click.echo(f"⏹️  Остановка демона (PID: {pid})...")
                
                # Отправляем SIGTERM
                os.kill(pid, signal.SIGTERM)
                
                # Ждем завершения до 30 секунд
                import time
                for _ in range(30):
                    if not psutil.pid_exists(pid):
                        break
                    time.sleep(1)
                
                # Если не завершился, принудительно убиваем
                if psutil.pid_exists(pid):
                    click.echo("⚡ Принудительная остановка...")
                    os.kill(pid, signal.SIGKILL)
                
                click.echo("✅ Демон остановлен")
            else:
                click.echo("❌ Процесс демона не найден")
            
            pid_file.unlink()
            
        except Exception as e:
            click.echo(f"❌ Ошибка остановки демона: {e}")
    
    elif action == 'status':
        try:
            from core.task_database import TaskDatabase
            db = TaskDatabase()
            
            # Сначала очищаем мертвые процессы
            db.cleanup_dead_processes()
            
            # Проверяем демон через БД
            daemon_pid = db.get_process_pid("scheduler_daemon")
            web_pid = db.get_process_pid("web_server")
            
            if daemon_pid and psutil.pid_exists(daemon_pid):
                process = psutil.Process(daemon_pid)
                click.echo(f"✅ Демон запущен")
                click.echo(f"   PID: {daemon_pid}")
                
                if web_pid and psutil.pid_exists(web_pid):
                    click.echo(f"   Веб-панель: PID {web_pid} (http://localhost:8000)")
                else:
                    click.echo(f"   Веб-панель: ❌ не запущена")
                click.echo(f"   CPU: {process.cpu_percent():.1f}%")
                click.echo(f"   Memory: {process.memory_info().rss / 1024 / 1024:.1f} MB")
                click.echo(f"   Started: {datetime.fromtimestamp(process.create_time()).strftime('%Y-%m-%d %H:%M:%S')}")
                
                # Показываем последние несколько строк общего лога
                log_path = Path('logs/app.log')
                if log_path.exists():
                    click.echo("\n📄 Последние записи лога:")
                    try:
                        lines = log_path.read_text(encoding='utf-8').strip().split('\n')
                        for line in lines[-5:]:
                            if line.strip():
                                click.echo(f"   {line}")
                    except:
                        click.echo("   (не удалось прочитать лог)")
            else:
                click.echo("❌ Демон не запущен (не найден в БД или процесс мертв)")
                if web_pid and psutil.pid_exists(web_pid):
                    click.echo(f"⚠️  Веб-панель работает отдельно: PID {web_pid}")
                    
        except Exception as e:
            click.echo(f"❌ Ошибка получения статуса: {e}")
    
    elif action == 'restart':
        click.echo("🔄 Перезапуск демона...")
        # Останавливаем
        ctx = click.get_current_context()
        ctx.invoke(daemon, action='stop', config=config, log_level=log_level, background=background)
        
        # Небольшая пауза
        import time
        time.sleep(2)
        
        # Запускаем
        ctx.invoke(daemon, action='start', config=config, log_level=log_level, background=background)


if __name__ == '__main__':
    cli()
