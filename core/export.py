"""
Оптимизированный экспортер вакансий в Excel
Базируется на лучших практиках из wh_excel_writer.py и wh_logger_config.py

Автор: AI Assistant (Senior Python Developer)
Дата: 20.09.2025 08:10:00
"""

import pandas as pd
import logging
import sqlite3
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from datetime import datetime
import json

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


logger = logging.getLogger(__name__)

# // Chg_EXPORT_FORMATS_2009: Определение форматов экспорта
EXPORT_FORMATS = {
    'brief': {
        'name': 'Краткий формат',
        'description': 'Основные поля для быстрого анализа',
        'columns': [
            'Название', 'Компания', 'Зарплата от', 'Зарплата до', 'Валюта',
            'Опыт', 'Город', 'Дата публикации', 'Ссылка', 'Фильтр'
        ],
        'sql_fields': [
            'title', 'company', 'salary_from', 'salary_to', 'currency',
            'experience', 'area', 'published_at', 'url', 'filter_id'
        ]
    },
    'full': {
        'name': 'Полный формат',
        'description': 'Все основные поля БД',
        'columns': [
            'ID', 'HH ID', 'Название', 'Компания', 'Компания ID',
            'Зарплата от', 'Зарплата до', 'Валюта', 'Опыт', 'График работы',
            'Занятость', 'Город', 'Ключевые навыки', 'Дата публикации',
            'Ссылка', 'Фильтр', 'Контент-хэш', 'Создано', 'Обновлено'
        ],
        'sql_fields': [
            'id', 'hh_id', 'title', 'company', 'employer_id',
            'salary_from', 'salary_to', 'currency', 'experience', 'schedule',
            'employment', 'area', 'key_skills', 'published_at',
            'url', 'filter_id', 'content_hash', 'created_at', 'updated_at'
        ]
    },
    'analytical': {
        'name': 'Аналитический формат',
        'description': 'С результатами плагинов и анализа',
        'columns': [
            'Название', 'Компания', 'Зарплата от', 'Зарплата до', 'Валюта',
            'Опыт', 'Город', 'Занятость', 'График',
            'Описание', 'Фильтр', 'Дата публикации', 'Ссылка'
        ],
        'sql_fields': [
            'title', 'company', 'salary_from', 'salary_to', 'currency',
            'experience', 'area', 'employment', 'schedule',
            'description', 'filter_id', 'published_at', 'url'
        ]
    }
}


class VacancyExporter:
    """Оптимизированный экспортер вакансий в Excel"""
    
    def __init__(self, db_path: str = "data/hh_v4.sqlite3"):
        self.db_path = db_path
        
        if not HAS_OPENPYXL:
            logger.error("openpyxl не установлен. Установите: pip install openpyxl")
            raise ImportError("openpyxl is required for Excel export")
    
    def export_to_excel(self, 
                       output_path: Union[str, Path],
                       format_type: str = 'brief',
                       limit: Optional[int] = None,
                       filters: Optional[Dict[str, Any]] = None,
                       include_description: bool = False) -> Dict[str, Any]:
        """
        Экспорт вакансий в Excel файл
        
        Args:
            output_path: Путь к выходному файлу
            format_type: Тип формата ('brief', 'full', 'analytical')
            limit: Максимальное количество записей (None = все)
            filters: Дополнительные фильтры для SQL запроса
            include_description: Включать ли описание вакансий (увеличивает размер)
            
        Returns:
            Dict с результатами экспорта (статистика, ошибки)
        """
        logger.info(f"🚀 Начало экспорта в формате '{format_type}' в файл: {output_path}")
        
        start_time = datetime.now()
        result = {
            'success': False,
            'file_path': str(output_path),
            'format_type': format_type,
            'records_exported': 0,
            'file_size_mb': 0,
            'export_time_seconds': 0,
            'errors': []
        }
        
        try:
            # Проверяем формат
            if format_type not in EXPORT_FORMATS:
                raise ValueError(f"Неизвестный формат: {format_type}. Доступные: {list(EXPORT_FORMATS.keys())}")
            
            format_config = EXPORT_FORMATS[format_type]
            
            # Получаем данные из БД
            data = self._fetch_vacancy_data(format_config, limit, filters, include_description)
            
            if not data:
                logger.warning("Нет данных для экспорта")
                result['errors'].append("Нет данных для экспорта")
                return result
            
            # Конвертируем в DataFrame
            df = self._convert_to_dataframe(data, format_config)
            
            # Экспортируем в Excel
            self._write_to_excel(df, output_path, format_config)
            
            # Собираем статистику
            output_file = Path(output_path)
            if output_file.exists():
                result.update({
                    'success': True,
                    'records_exported': len(df),
                    'file_size_mb': round(output_file.stat().st_size / (1024 * 1024), 2),
                    'export_time_seconds': round((datetime.now() - start_time).total_seconds(), 2)
                })
                
                logger.info(f"✅ Экспорт завершен: {result['records_exported']} записей, "
                           f"{result['file_size_mb']} МБ, {result['export_time_seconds']} сек")
            
        except Exception as e:
            logger.error(f"❌ Ошибка экспорта: {e}", exc_info=True)
            result['errors'].append(str(e))
        
        return result
    
    def _fetch_vacancy_data(self, 
                          format_config: Dict[str, Any], 
                          limit: Optional[int] = None,
                          filters: Optional[Dict[str, Any]] = None,
                          include_description: bool = False) -> List[Dict[str, Any]]:
        """Получение данных вакансий из БД с оптимизацией"""
        
        # Базовые поля
        sql_fields = format_config['sql_fields'].copy()
        
        # Добавляем описание если нужно
        if include_description and 'description' not in sql_fields:
            sql_fields.append('description')
        
        # Формируем SQL запрос
        fields_str = ', '.join(sql_fields)
        base_query = f"SELECT {fields_str} FROM vacancies"
        
        # Добавляем фильтры
        where_conditions = []
        params = []
        
        if filters:
            if 'date_from' in filters:
                where_conditions.append("created_at >= ?")
                params.append(filters['date_from'])
            if 'date_to' in filters:
                where_conditions.append("created_at <= ?")
                params.append(filters['date_to'])
            if 'min_salary' in filters:
                where_conditions.append("salary_from >= ?")
                params.append(filters['min_salary'])
            if 'area_name' in filters:
                where_conditions.append("area LIKE ?")
                params.append(f"%{filters['area_name']}%")
        
        # Собираем финальный запрос
        if where_conditions:
            base_query += " WHERE " + " AND ".join(where_conditions)
        
        base_query += " ORDER BY created_at DESC"
        
        if limit:
            base_query += f" LIMIT {limit}"
        
        logger.debug(f"SQL запрос: {base_query}")
        logger.debug(f"Параметры: {params}")
        
        # Выполняем запрос
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row  # Для доступа к колонкам по имени
                cursor = conn.execute(base_query, params)
                rows = cursor.fetchall()
                
                # Конвертируем в список словарей
                data = [dict(row) for row in rows]
                
                logger.info(f"📊 Получено {len(data)} записей из БД")
                return data
                
        except Exception as e:
            logger.error(f"Ошибка выполнения SQL запроса: {e}")
            raise
    
    def _convert_to_dataframe(self, data: List[Dict[str, Any]], format_config: Dict[str, Any]) -> pd.DataFrame:
        """Конвертация данных в DataFrame с оптимизацией"""
        
        if not data:
            return pd.DataFrame()
        
        # Создаем DataFrame
        df = pd.DataFrame(data)
        
        # Переименовываем колонки согласно формату
        if len(format_config['columns']) == len(format_config['sql_fields']):
            column_mapping = dict(zip(format_config['sql_fields'], format_config['columns']))
            df = df.rename(columns=column_mapping)
        
        # Обработка специальных полей
        for col in df.columns:
            if 'Дата' in col and col in df.columns:
                # Форматируем даты
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d.%m.%Y %H:%M')
            
            elif 'Ключевые навыки' in col and col in df.columns:
                # Обрабатываем JSON навыки
                df[col] = df[col].apply(self._format_skills)
            
            elif 'Зарплата' in col and col in df.columns:
                # Форматируем зарплаты
                df[col] = df[col].apply(lambda x: f"{int(x):,}".replace(',', ' ') if pd.notna(x) and x > 0 else '')
        
        # Добавляем колонку "Статус" если её нет
        if 'Статус' not in df.columns:
            df['Статус'] = ''
        
        logger.debug(f"DataFrame создан: {df.shape[0]} строк, {df.shape[1]} колонок")
        return df
    
    def _format_skills(self, skills_json: str) -> str:
        """Форматирование навыков из JSON"""
        if not skills_json:
            return ''
        
        try:
            if isinstance(skills_json, str):
                skills_list = json.loads(skills_json)
            else:
                skills_list = skills_json
            
            if isinstance(skills_list, list):
                return ', '.join(skills_list[:10])  # Ограничиваем 10 навыками
            else:
                return str(skills_list)[:100]  # Ограничиваем длину
                
        except (json.JSONDecodeError, TypeError):
            return str(skills_json)[:100] if skills_json else ''
    
    def _write_to_excel(self, df: pd.DataFrame, output_path: Union[str, Path], format_config: Dict[str, Any]):
        """Запись DataFrame в оптимизированный Excel файл"""
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Создаем стандартный ExcelWriter (без неподдерживаемых options)
        with pd.ExcelWriter(
            output_path,
            engine='openpyxl'
        ) as writer:
            
            # Записываем основной лист
            sheet_name = f"Вакансии_{format_config['name'].replace(' ', '_')}"
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Применяем форматирование
            worksheet = writer.sheets[sheet_name]
            self._format_worksheet(worksheet, format_config)
            
            # Добавляем лист с информацией об экспорте
            self._add_info_sheet(writer, df, format_config)
        
        logger.info(f"📁 Файл Excel сохранен: {output_path}")
    
    def _format_worksheet(self, worksheet, format_config: Dict[str, Any]):
        """
        Форматирование листа Excel (на основе format_worksheet из wh_logger_config.py)
        """
        # Добавляем автофильтр и закрепляем первую строку
        if worksheet.dimensions:
            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.freeze_panes = 'A2'
        
        # Создаем стили
        base_alignment = Alignment(horizontal='left', vertical='top', wrap_text=False)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Применяем стили и рассчитываем ширину колонок
        max_lengths = [0] * (worksheet.max_column or 1)
        
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value:
                    # Применяем базовый стиль
                    cell.alignment = base_alignment
                    
                    if cell.row == 1:  # Заголовки
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Рассчитываем ширину колонки
                    try:
                        cell_length = len(str(cell.value))
                        if cell.column <= len(max_lengths):
                            max_lengths[cell.column - 1] = max(
                                max_lengths[cell.column - 1], 
                                min(cell_length, 50)  # Максимум 50 символов
                            )
                    except (IndexError, TypeError):
                        pass
        
        # Устанавливаем ширину колонок (от 10 до 40 символов)
        for col, max_length in enumerate(max_lengths, 1):
            adjusted_width = max(min(max_length + 2, 40), 10)
            try:
                worksheet.column_dimensions[get_column_letter(col)].width = adjusted_width
            except Exception:
                pass
        
        logger.debug(f"Форматирование листа завершено: {worksheet.max_row} строк, {worksheet.max_column} колонок")
    
    def _add_info_sheet(self, writer, df: pd.DataFrame, format_config: Dict[str, Any]):
        """Добавление информационного листа"""
        
        info_data = {
            'Параметр': [
                'Формат экспорта',
                'Описание формата', 
                'Количество записей',
                'Количество колонок',
                'Дата экспорта',
                'Время экспорта',
                'Путь к БД'
            ],
            'Значение': [
                format_config['name'],
                format_config['description'],
                len(df),
                len(df.columns),
                datetime.now().strftime('%d.%m.%Y %H:%M:%S'),
                datetime.now().strftime('%H:%M:%S'),
                self.db_path
            ]
        }
        
        info_df = pd.DataFrame(info_data)
        info_df.to_excel(writer, sheet_name='Информация', index=False)
        
        # Форматируем информационный лист
        info_sheet = writer.sheets['Информация']
        info_sheet.column_dimensions['A'].width = 25
        info_sheet.column_dimensions['B'].width = 50
        
        # Заголовки
        for cell in info_sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    def get_export_formats(self) -> Dict[str, Dict[str, Any]]:
        """Получение доступных форматов экспорта"""
        return EXPORT_FORMATS.copy()
    
    def get_vacancy_count(self, filters: Optional[Dict[str, Any]] = None) -> int:
        """Получение количества вакансий для экспорта"""
        base_query = "SELECT COUNT(*) FROM vacancies"
        
        where_conditions = []
        params = []
        
        if filters:
            if 'date_from' in filters:
                where_conditions.append("created_at >= ?")
                params.append(filters['date_from'])
            if 'date_to' in filters:
                where_conditions.append("created_at <= ?")
                params.append(filters['date_to'])
            if 'min_salary' in filters:
                where_conditions.append("salary_from >= ?")
                params.append(filters['min_salary'])
        
        if where_conditions:
            base_query += " WHERE " + " AND ".join(where_conditions)
        
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.execute(base_query, params)
                count = cursor.fetchone()[0]
                return count
        except Exception as e:
            logger.error(f"Ошибка подсчета вакансий: {e}")
            return 0


# // Chg_EXPORT_HELPER_2009: Вспомогательные функции для быстрого экспорта
def quick_export(output_path: Union[str, Path], 
                format_type: str = 'brief',
                limit: int = 1000,
                db_path: str = "data/hh_v4.sqlite3") -> Dict[str, Any]:
    """Быстрый экспорт с параметрами по умолчанию"""
    exporter = VacancyExporter(db_path)
    return exporter.export_to_excel(output_path, format_type, limit)


def export_with_filters(output_path: Union[str, Path],
                       date_from: Optional[str] = None,
                       min_salary: Optional[int] = None,
                       area_name: Optional[str] = None,
                       format_type: str = 'brief') -> Dict[str, Any]:
    """Экспорт с фильтрами"""
    filters = {}
    if date_from:
        filters['date_from'] = date_from
    if min_salary:
        filters['min_salary'] = min_salary
    if area_name:
        filters['area_name'] = area_name
    
    exporter = VacancyExporter()
    return exporter.export_to_excel(output_path, format_type, filters=filters)
